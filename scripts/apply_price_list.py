import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "정보서류" / "타일앤바스플러스단가표.xlsx"
DB = ROOT / "data" / "products.json"
PRODUCTS_JS = ROOT / "products-db.js"
REPORT = ROOT / "data" / "price-match-report.json"

TILE_CATEGORIES = {"벽 타일", "바닥타일", "포세린", "포인트타일", "수조타일"}


def normalize_size(value):
    text = str(value or "").upper().replace("×", "*").replace("X", "*")
    text = re.sub(r"\s+", "", text)
    return text


def normalize_text(value):
    return re.sub(r"\s+", "", str(value or "").lower())


def read_price_rows():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["품목등록"]
    headers = [cell.value for cell in ws[2]]
    rows = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        item = dict(zip(headers, row))
        name = str(item.get("품목명") or "").strip()
        if not name:
            continue

        item["_name_norm"] = normalize_text(name)
        item["_size_norm"] = normalize_size(item.get("규격정보"))
        item["_category"] = str(item.get("분류명") or "").strip()
        item["_codes"] = set(re.findall(r"\b[A-Z]?\d{3,5}\b", name))
        rows.append(item)

    return rows


def row_has_price(row):
    return row.get("입고단가") is not None or row.get("출고단가") is not None


def is_tile_price_row(row):
    return row["_category"] in TILE_CATEGORIES


def choose_price_row(product, rows):
    code = str(product.get("catalogCode") or "").strip()
    size = normalize_size(product.get("size"))
    product_name = normalize_text(product.get("name"))

    candidates = []

    if code and size:
        candidates = [
            row for row in rows
            if code in row["_codes"]
            and row["_size_norm"] == size
            and is_tile_price_row(row)
            and row_has_price(row)
        ]

    if not candidates and size and product.get("productType") == "tile":
        # Fallback for hand-entered tile products that may not have catalog codes.
        candidates = [
            row for row in rows
            if row["_size_norm"] == size
            and is_tile_price_row(row)
            and row_has_price(row)
            and (normalize_text(product.get("maker")) in row["_name_norm"] or row["_name_norm"] in product_name)
        ]

    if not candidates:
        return None

    # Prefer rows with both prices, then the longest product name for specificity.
    candidates.sort(
        key=lambda row: (
            row.get("입고단가") is not None and row.get("출고단가") is not None,
            len(str(row.get("품목명") or "")),
        ),
        reverse=True,
    )
    return candidates[0]


def apply_price(product, row):
    if product.get("priceUpdatedBy") == "scripts/apply_price_list.py":
        product["retailPrice"] = 0

    before = {
        "costPrice": product.get("costPrice", 0),
        "retailPrice": product.get("retailPrice", 0),
        "wholesalePrice": product.get("wholesalePrice", 0),
    }

    incoming = row.get("입고단가")
    outgoing = row.get("출고단가")

    if incoming is not None:
        product["costPrice"] = int(round(float(incoming)))
    if outgoing is not None:
        product["wholesalePrice"] = int(round(float(outgoing)))

    product["priceSource"] = "타일앤바스플러스단가표.xlsx"
    product["priceSourceCode"] = str(row.get("품목코드") or "")
    product["priceSourceName"] = str(row.get("품목명") or "")
    product["priceUpdatedBy"] = "scripts/apply_price_list.py"

    after = {
        "costPrice": product.get("costPrice", 0),
        "retailPrice": product.get("retailPrice", 0),
        "wholesalePrice": product.get("wholesalePrice", 0),
    }
    return before, after


def main():
    products = json.loads(DB.read_text(encoding="utf-8"))
    rows = read_price_rows()
    report = {
        "excel": str(EXCEL),
        "db": str(DB),
        "products_total": len(products),
        "excel_rows": len(rows),
        "matched": [],
        "unmatched_catalog_products": [],
    }

    for product in products:
        row = choose_price_row(product, rows)
        if row:
            before, after = apply_price(product, row)
            report["matched"].append({
                "id": product["id"],
                "name": product.get("name"),
                "size": product.get("size"),
                "catalogCode": product.get("catalogCode"),
                "priceRowCode": row.get("품목코드"),
                "priceRowName": row.get("품목명"),
                "priceRowSize": row.get("규격정보"),
                "priceRowCategory": row.get("분류명"),
                "before": before,
                "after": after,
            })
        elif product.get("catalogCode"):
            report["unmatched_catalog_products"].append({
                "id": product["id"],
                "name": product.get("name"),
                "size": product.get("size"),
                "catalogCode": product.get("catalogCode"),
            })

    DB.write_text(json.dumps(products, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    PRODUCTS_JS.write_text("window.PRODUCTS_DB = " + json.dumps(products, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({
        "products_total": len(products),
        "excel_rows": len(rows),
        "matched": len(report["matched"]),
        "unmatched_catalog_products": len(report["unmatched_catalog_products"]),
        "report": str(REPORT),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

import collections
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "정보서류" / "타일앤바스플러스단가표.xlsx"
DB = ROOT / "data" / "products.json"

products = json.loads(DB.read_text(encoding="utf-8"))
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["품목등록"]
headers = [cell.value for cell in ws[2]]
rows = []

for row in ws.iter_rows(min_row=3, values_only=True):
    item = dict(zip(headers, row))
    if item.get("품목명"):
        rows.append(item)

print("excel rows", len(rows), "db products", len(products))
print("categories", collections.Counter(str(row.get("분류명")) for row in rows).most_common(30))

code_to_rows = collections.defaultdict(list)
for row in rows:
    text = str(row.get("품목명") or "")
    for code in re.findall(r"\b[A-Z]?\d{3,5}\b", text):
        code_to_rows[code].append(row)

matched = []
for product in products:
    code = product.get("catalogCode")
    if code and code in code_to_rows:
        row = code_to_rows[code][0]
        matched.append((
            product["id"],
            code,
            len(code_to_rows[code]),
            row.get("품목명"),
            row.get("입고단가"),
            row.get("출고단가"),
        ))

print("catalog code matched", len(matched))
print("sample")
for item in matched[:30]:
    print(item)

print("tile generic price rows")
for row in rows[:120]:
    if row.get("규격정보") or row.get("분류명") in ["벽 타일", "바닥타일", "포세린"]:
        print(row.get("품목코드"), row.get("품목명"), row.get("규격정보"), row.get("분류명"), row.get("입고단가"), row.get("출고단가"))

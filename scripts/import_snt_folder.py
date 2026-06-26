from __future__ import annotations

import json
import math
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook
from PIL import Image, ImageStat


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "incoming" / "price-image-db" / "SNT타일"
WORKBOOK_PATH = SOURCE_DIR / "SNT 타일 (1).xlsx"
PRODUCTS_PATH = ROOT / "data" / "products.json"
OUTPUT_DIR = ROOT / "outputs" / "snt-import"
BRAND = "SNT"

TIMESTAMP = datetime.now().strftime("%Y%m%d-%H%M%S")


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"SNT 엑셀 파일을 찾지 못했습니다: {WORKBOOK_PATH}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    existing_products = read_json_array(PRODUCTS_PATH)
    workbook_rows = read_workbook_rows(WORKBOOK_PATH)
    image_index = build_image_index(SOURCE_DIR)

    products, report_rows = build_snt_products(workbook_rows, image_index)
    products.extend(build_extra_image_products(workbook_rows, image_index))
    products.sort(key=lambda item: (item["catalogPage"], item["managementCode"]))

    backup_path = PRODUCTS_PATH.with_name(f"products.backup-before-snt-import-{TIMESTAMP}.json")
    shutil.copyfile(PRODUCTS_PATH, backup_path)

    without_snt = [
        product
        for product in existing_products
        if not is_snt_product(product)
    ]
    final_products = without_snt + products
    write_json(PRODUCTS_PATH, final_products)

    report = {
        "ok": True,
        "brand": BRAND,
        "sourceDir": str(SOURCE_DIR),
        "workbookPath": str(WORKBOOK_PATH),
        "backupPath": str(backup_path),
        "productsPath": str(PRODUCTS_PATH),
        "beforeProductCount": len(existing_products),
        "removedExistingSntCount": len(existing_products) - len(without_snt),
        "importedCount": len(products),
        "finalProductCount": len(final_products),
        "withImages": sum(1 for product in products if product.get("image")),
        "withoutImages": sum(1 for product in products if not product.get("image")),
        "withCostPrice": sum(1 for product in products if product.get("costPrice")),
        "withoutCostPrice": sum(1 for product in products if not product.get("costPrice")),
        "bySourceCategory": dict(Counter(product.get("sourceCategoryName", "") for product in products)),
        "byFinish": dict(Counter(product.get("finish", "") or "마감 미확인" for product in products)),
        "byMaterial": dict(Counter(product.get("material", "") or "재질 미확인" for product in products)),
        "imageMissingRows": [row for row in report_rows if not row.get("image")],
        "priceMissingRows": [row for row in report_rows if not row.get("costPrice")],
    }
    report_path = OUTPUT_DIR / f"snt-import-report-{TIMESTAMP}.json"
    imported_path = OUTPUT_DIR / f"snt-products-{TIMESTAMP}.json"
    write_json(report_path, report)
    write_json(imported_path, products)

    print(json.dumps({
        "ok": True,
        "backupPath": str(backup_path),
        "reportPath": str(report_path),
        "importedPath": str(imported_path),
        "importedCount": len(products),
        "withImages": report["withImages"],
        "withoutImages": report["withoutImages"],
        "withoutCostPrice": report["withoutCostPrice"],
        "finalProductCount": len(final_products),
    }, ensure_ascii=False, indent=2))


def read_workbook_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    rows: list[dict] = []
    sequence = 0
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        headers = [str(value).strip() if value is not None else "" for value in values[0]]
        for row_index, values_row in enumerate(values[1:], start=2):
            if not any(value is not None and str(value).strip() for value in values_row):
                continue
            raw = dict(zip(headers, values_row))
            code = clean_cell(raw.get("품번"))
            size = normalize_size(clean_cell(raw.get("규격")))
            tile_type = clean_cell(raw.get("유형"))
            if not code:
                continue
            sequence += 1
            rows.append({
                "sequence": sequence,
                "sheet": sheet.title,
                "rowIndex": row_index,
                "code": code,
                "size": size,
                "type": tile_type,
            })
    return rows


def build_image_index(source_dir: Path) -> list[dict]:
    images: list[dict] = []
    for path in sorted(source_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".jpg", ".jpeg", ".png"}:
            continue
        if path.name.startswith("KakaoTalk_"):
            continue
        relative = path.relative_to(ROOT).as_posix()
        images.append({
            "path": path,
            "relative": relative,
            "url": "/" + quote(relative, safe="/-_.~()"),
            "folder": path.parent.name,
            "stem": path.stem,
            "codes": image_code_variants(path.stem),
        })
    return images


def build_snt_products(rows: list[dict], images: list[dict]) -> tuple[list[dict], list[dict]]:
    products: list[dict] = []
    report_rows: list[dict] = []
    used_ids: set[str] = set()

    for row in rows:
        image = find_image_for_row(row, images)
        price = price_for_row(row)
        size = row["size"]
        width, height = parse_size(size)
        thickness = parse_thickness(row["type"])
        material = infer_material(row)
        finish = infer_finish(row)
        color = infer_color(row, image)
        pattern = infer_pattern(row)
        sqm = sqm_per_box(row)
        pcs = pcs_per_box(row, sqm, width, height)
        unit = unit_label(row, pcs, sqm)
        source_category = source_category_name(row)
        code = normalized_display_code(row["code"], row)
        product_id = unique_id(f"snt-{slugify(code)}", used_ids)
        image_url = image["url"] if image else ""

        features = " / ".join(
            part for part in [
                source_category,
                row["type"],
                f"{thickness}T" if thickness else "",
                "단가표 VAT 포함",
                "재고 미제공",
                "이미지 미확인" if not image_url else "",
            ]
            if part
        )

        product = {
            "id": product_id,
            "managementCode": f"SNT-{code}",
            "majorCategory": BRAND,
            "productType": "tile",
            "kind": BRAND,
            "option": source_category,
            "name": product_name(code, size, row["type"]),
            "modelName": code,
            "size": size,
            "material": material,
            "patternCategory": pattern,
            "finish": finish,
            "surface": finish,
            "countryOfOrigin": "",
            "maker": BRAND,
            "unit": unit,
            "pcsPerBox": pcs,
            "sqmPerBox": sqm,
            "color": color,
            "features": features,
            "costPrice": price,
            "retailPrice": 0,
            "wholesalePrice": 0,
            "gradeAPrice": "",
            "gradeBPrice": "",
            "gradeCPrice": "",
            "stockQty": 0,
            "stockText": "재고 미제공",
            "image": image_url,
            "imageUrls": [image_url] if image_url else [],
            "originalImage": image_url,
            "closeImage": image_url,
            "detailImage": image_url,
            "daylightImage": "",
            "fluorescentImage": "",
            "sceneImage": "",
            "sourceSite": "SNT local price sheet",
            "sourceUrl": str(WORKBOOK_PATH),
            "sourceProductId": code,
            "sourceCategoryCode": row["sheet"],
            "sourceCategoryName": source_category,
            "catalogSource": BRAND,
            "catalogPage": row["sequence"],
            "lastSyncedAt": datetime.now().isoformat(timespec="milliseconds"),
            "sntImportSource": "folder-excel-price-image",
            "sntImportWorkbookSheet": row["sheet"],
            "sntImportWorkbookRow": row["rowIndex"],
            "sntImportImagePath": image["relative"] if image else "",
            "sntImportPriceRule": price_rule_for_row(row),
        }
        products.append(product)
        report_rows.append({
            "code": code,
            "sheet": row["sheet"],
            "rowIndex": row["rowIndex"],
            "size": size,
            "type": row["type"],
            "image": image["relative"] if image else "",
            "costPrice": price,
            "material": material,
            "finish": finish,
            "color": color,
        })

    return products, report_rows


def build_extra_image_products(rows: list[dict], images: list[dict]) -> list[dict]:
    row_codes = set()
    for row in rows:
        row_codes |= code_variants(row["code"])

    extras: list[dict] = []
    used_ids: set[str] = set()
    for image in images:
        if "1622" not in image["codes"]:
            continue
        if "1622" in row_codes:
            continue
        row = {
            "sequence": 60000,
            "sheet": "600X600 포슬레인",
            "rowIndex": 0,
            "code": "1622",
            "size": "600*600",
            "type": "포슬레인(그래뉼,펄)",
        }
        product_id = unique_id("snt-1622", used_ids)
        image_url = image["url"]
        product = {
            "id": product_id,
            "managementCode": "SNT-1622",
            "majorCategory": BRAND,
            "productType": "tile",
            "kind": BRAND,
            "option": source_category_name(row),
            "name": product_name("1622", "600*600", row["type"]),
            "modelName": "1622",
            "size": "600*600",
            "material": "포세린",
            "patternCategory": "패턴",
            "finish": "무광",
            "surface": "무광",
            "countryOfOrigin": "",
            "maker": BRAND,
            "unit": unit_label(row, 4, 1.44),
            "pcsPerBox": 4,
            "sqmPerBox": 1.44,
            "color": infer_color(row, image),
            "features": "600X600 포쉐린 / 포슬레인(그래뉼,펄) / 엑셀 미기재 이미지 품번 / 단가표 VAT 포함 / 재고 미제공",
            "costPrice": 12750,
            "retailPrice": 0,
            "wholesalePrice": 0,
            "gradeAPrice": "",
            "gradeBPrice": "",
            "gradeCPrice": "",
            "stockQty": 0,
            "stockText": "재고 미제공",
            "image": image_url,
            "imageUrls": [image_url],
            "originalImage": image_url,
            "closeImage": image_url,
            "detailImage": image_url,
            "daylightImage": "",
            "fluorescentImage": "",
            "sceneImage": "",
            "sourceSite": "SNT local image folder",
            "sourceUrl": str(image["path"]),
            "sourceProductId": "1622",
            "sourceCategoryCode": row["sheet"],
            "sourceCategoryName": source_category_name(row),
            "catalogSource": BRAND,
            "catalogPage": row["sequence"],
            "lastSyncedAt": datetime.now().isoformat(timespec="milliseconds"),
            "sntImportSource": "folder-image-extra",
            "sntImportImagePath": image["relative"],
            "sntImportPriceRule": "600X600 포쉐린 1622~1627",
        }
        extras.append(product)
    return extras


def find_image_for_row(row: dict, images: list[dict]) -> dict | None:
    variants = code_variants(row["code"])
    if "폼세라" in row["sheet"] or "폼세라" in row["type"]:
        variants |= {"폼세라", "폼세라믹"}

    matches = [image for image in images if variants & image["codes"]]
    if not matches:
        return None
    same_folder = [
        image for image in matches
        if folder_key(image["folder"]) == folder_key(row["sheet"])
    ]
    return (same_folder or matches)[0]


def source_category_name(row: dict) -> str:
    sheet = row["sheet"]
    tile_type = row["type"]
    if "100X300" in sheet:
        return "특수규격 100X300"
    if "250X400" in sheet:
        return "벽타일 250X400"
    if "300X600" in sheet:
        if any(token in tile_type for token in ["MZ", "민자파스텔", "액자", "템바"]):
            return "벽타일 300X600 특수"
        return "벽타일 300X600"
    if "200X200" in sheet:
        if "석재" in tile_type:
            return "석재 200X200"
        if "포슬레인" in tile_type:
            return "바닥타일 200X200 패턴"
        return "바닥타일 200X200"
    if "300X300" in sheet:
        return "바닥타일 300X300"
    if "600X600 폴리싱" in sheet:
        return "폴리싱 600X600"
    if "600X600" in sheet:
        return "포쉐린 600X600"
    if "600X1200" in sheet:
        return "포쉐린 600X1200"
    if "800X800" in sheet:
        return "포쉐린 800X800"
    if "계단" in sheet:
        return "계단타일"
    if "점자" in sheet:
        return "점자타일"
    if "폼세라" in sheet:
        return "폼세라"
    return sheet


def price_for_row(row: dict) -> int:
    sheet = row["sheet"]
    code = normalized_code(row["code"])
    tile_type = row["type"]

    if "100X300" in sheet:
        return 11500 if code.endswith("3101") else 12000
    if "250X400" in sheet:
        return 8650 + (200 if "무광" in tile_type else 0)
    if "300X600" in sheet:
        if any(token in tile_type for token in ["MZ", "민자파스텔", "액자", "템바"]):
            return 9500 + (1000 if infer_text_color(row) not in {"화이트", ""} else 0)
        base = 8450
        if any(token in tile_type for token in ["테라조", "캔디"]):
            base += 300
        if "바둑판" in tile_type:
            base += 1000
        return base
    if "200X200" in sheet:
        if "석재" in tile_type:
            return 13500
        if "포슬레인" in tile_type:
            return 14000
        return 8650
    if "300X300" in sheet:
        return 8950 if any(token in tile_type for token in ["까라라", "테라조"]) else 8450
    if "600X600 폴리싱" in sheet:
        return 14250 if code == "9601" else 13050
    if "600X600" in sheet:
        numeric = number_part(code)
        if numeric in {6657, 6658, 6659, 6660, 6661, 6662, 6663}:
            return 11150
        if numeric in {6614, 6615, 6616, 6617, 6100, 6664, 6665}:
            return 11700
        if numeric in {6655, 6656}:
            return 12200
        if numeric in {6651, 6652, 6653, 6654, 6070, 1622, 1623, 1624, 1625, 1626, 1627}:
            return 12750
        return 10650
    if "600X1200" in sheet:
        numeric = number_part(code)
        if numeric == 1203:
            return 18250
        if numeric >= 1251:
            return 17750
        return 16250
    if "800X800" in sheet:
        return 27000
    if "계단" in sheet:
        return 9000 if row["size"] == "200*300" else 12500
    if "점자" in sheet:
        return 2650
    return 0


def price_rule_for_row(row: dict) -> str:
    return f"{source_category_name(row)} / {row['type']}".strip(" /")


def sqm_per_box(row: dict) -> float | None:
    sheet = row["sheet"]
    tile_type = row["type"]
    if "100X300" in sheet:
        return 0.99
    if "250X400" in sheet:
        return 1.50
    if "300X600" in sheet:
        return 1.44
    if "200X200" in sheet:
        if "석재" in tile_type:
            return 0.88
        if "포슬레인" in tile_type:
            return 1.0
        return 1.48
    if "300X300" in sheet:
        return 1.44
    if "600X600" in sheet:
        return 1.44
    if "600X1200" in sheet:
        return 1.44
    if "800X800" in sheet:
        return 1.92
    if "계단" in sheet:
        return 0.96 if row["size"] == "200*300" else 1.44
    if "점자" in sheet:
        return 0.72
    return None


def pcs_per_box(row: dict, sqm: float | None, width: int | None, height: int | None) -> int | None:
    if "점자" in row["sheet"]:
        return 8
    if sqm is None or not width or not height:
        return None
    area = (width / 1000) * (height / 1000)
    if not area:
        return None
    return int(round(sqm / area))


def unit_label(row: dict, pcs: int | None, sqm: float | None) -> str:
    if "점자" in row["sheet"]:
        return "장"
    parts = ["BOX"]
    if pcs:
        parts.append(f"들이({pcs})")
    if sqm:
        parts.append(f"{format_decimal(sqm)}㎡")
    return "/".join(parts)


def infer_material(row: dict) -> str:
    sheet = row["sheet"]
    tile_type = row["type"]
    size = row["size"]
    width, height = parse_size(size)
    max_side = max(width or 0, height or 0)
    min_side = min(width or 0, height or 0)
    text = f"{sheet} {tile_type}"

    if "폼세라" in text:
        return "폼세라믹"
    if "석재" in text:
        return "석재타일"
    if "포슬레인" in text or "포쉐린" in text or "폴리싱" in text:
        return "포세린"
    if "점자" in text:
        return "자기질"
    if "계단" in text:
        return "자기질"
    if max_side >= 600 and min_side >= 600:
        return "포세린"
    if "100X300" in sheet or "250X400" in sheet or "300X600" in sheet:
        return "도기질"
    if "200X200" in sheet or "300X300" in sheet:
        return "자기질"
    return ""


def infer_finish(row: dict) -> str:
    text = f"{row['sheet']} {row['type']}"
    if "폴리싱" in text or "유광" in text:
        return "유광"
    if "반무광" in text:
        return "반무광"
    if "폼세라" in text:
        return ""
    return "무광"


def infer_pattern(row: dict) -> str:
    text = normalize_text(f"{row['sheet']} {row['type']} {row['code']}")
    if any(token in text for token in ["테라조", "캔디", "그래뉼", "점자"]):
        return "테라조" if "테라조" in text else "패턴"
    if any(token in text for token in ["까라라", "카라라", "폴리싱"]):
        return "마블"
    if any(token in text for token in ["석재", "포슬레인", "포쉐린", "스톤"]):
        return "스톤"
    if any(token in text for token in ["바둑판", "MZ", "액자", "템바", "금형", "씽킹"]):
        return "패턴"
    if "폼세라" in text:
        return "기타"
    return "솔리드"


def infer_color(row: dict, image: dict | None) -> str:
    text_color = infer_text_color(row)
    if text_color:
        return text_color
    if image:
        try:
            return classify_image_color(image["path"])
        except Exception:
            return ""
    return ""


def infer_text_color(row: dict) -> str:
    text = normalize_text(f"{row.get('code', '')} {row.get('type', '')}")
    mappings = [
        (["백색", "화이트", "WHITE"], "화이트"),
        (["연회색", "라이트그레이"], "라이트그레이"),
        (["회색", "그레이", "GREY", "GRAY"], "그레이"),
        (["아이보리", "IVORY"], "아이보리"),
        (["베이지", "BEIGE"], "베이지"),
        (["블랙", "BLACK", "검정"], "블랙"),
        (["브라운", "BROWN"], "브라운"),
        (["그린", "GREEN"], "그린"),
        (["블루", "BLUE"], "블루"),
        (["핑크", "PINK"], "핑크"),
        (["옐로우", "YELLOW"], "옐로우"),
    ]
    for needles, color in mappings:
        if any(normalize_text(needle) in text for needle in needles):
            return color
    return ""


def classify_image_color(path: Path) -> str:
    with Image.open(path) as image:
        image = image.convert("RGB")
        width, height = image.size
        crop_box = (
            max(0, int(width * 0.08)),
            max(0, int(height * 0.08)),
            min(width, int(width * 0.92)),
            min(height, int(height * 0.92)),
        )
        image = image.crop(crop_box)
        image.thumbnail((96, 96))
        stat = ImageStat.Stat(image)
        r, g, b = stat.mean[:3]
    return classify_rgb(r, g, b)


def classify_rgb(r: float, g: float, b: float) -> str:
    brightness = (r + g + b) / 3
    spread = max(r, g, b) - min(r, g, b)
    if brightness < 55:
        return "블랙"
    if spread < 12:
        if brightness > 225:
            return "화이트"
        if brightness > 175:
            return "라이트그레이"
        if brightness > 95:
            return "그레이"
        return "다크그레이"
    if r > g + 22 and r > b + 22:
        if g > b + 20 and r > 120:
            return "베이지" if brightness > 145 else "브라운"
        return "핑크" if brightness > 145 else "브라운"
    if g > r + 18 and g > b + 18:
        return "그린"
    if b > r + 18 and b > g + 18:
        return "블루"
    if r > 150 and g > 135 and b < 125:
        return "베이지"
    if r > 120 and g > 105 and b > 85:
        return "아이보리" if brightness > 175 else "베이지"
    return "베이지" if brightness > 130 else "브라운"


def normalized_display_code(code: str, row: dict) -> str:
    cleaned = str(code).strip()
    if "폼세라" in row["sheet"] or "폼세라" in row["type"]:
        thickness = parse_thickness(row["type"])
        return f"FOM-CERA-{thickness}T" if thickness else "FOM-CERA"
    return re.sub(r"\s+", " ", cleaned).upper().replace(" ", "-")


def normalized_code(code: str) -> str:
    return re.sub(r"[^0-9A-Z-]+", "", str(code).upper())


def number_part(code: str) -> int:
    match = re.search(r"\d+", str(code))
    return int(match.group(0)) if match else 0


def code_variants(code: str) -> set[str]:
    key = normalized_code(code)
    variants = {key}
    if key.startswith("SNT"):
        variants.add(key[3:])
    else:
        variants.add("SNT" + key)
    if key.endswith("-S"):
        variants.add(key[:-2])
    variants.add(key.replace("-S", ""))
    for number in re.findall(r"\d+", key):
        variants.add(number)
    return {variant for variant in variants if variant}


def image_code_variants(stem: str) -> set[str]:
    compact = stem.replace(" ", "")
    if "폼세라" in compact:
        return {"폼세라", "폼세라믹"}
    variants: set[str] = set()
    for part in re.split(r"[,/&+\s]+", stem.upper().replace("_", " ")):
        part = part.strip(" .()[]{}")
        if not part:
            continue
        match = re.match(r"(SNT)?\s*([A-Z]*\d+[A-Z-]*\d*[A-Z-]*)", part)
        if match:
            code = normalized_code((match.group(1) or "") + match.group(2))
            variants.add(code)
            if code.startswith("SNT"):
                variants.add(code[3:])
            if code.endswith("-S"):
                variants.add(code[:-2])
            variants.add(code.replace("-S", ""))
        for number in re.findall(r"\d+", part):
            variants.add(number)
            if len(number) >= 5:
                variants.add(number[-4:])
    return {variant for variant in variants if variant}


def folder_key(value: str) -> str:
    return re.sub(r"[^0-9A-Z가-힣]", "", str(value).upper()).replace("포슬레인", "").replace("폴리싱", "").replace("반무광", "")


def product_name(code: str, size: str, tile_type: str) -> str:
    return " ".join(part for part in [f"SNT {code}", size, tile_type] if part)


def unique_id(base: str, used_ids: set[str]) -> str:
    candidate = base
    index = 2
    while candidate in used_ids:
        candidate = f"{base}-{index}"
        index += 1
    used_ids.add(candidate)
    return candidate


def slugify(value: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z가-힣]+", "-", str(value).strip()).strip("-").lower()
    return slug or "unknown"


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_size(value: str) -> str:
    return re.sub(r"\s*[xX×]\s*", "*", value.strip())


def parse_size(value: str) -> tuple[int | None, int | None]:
    match = re.search(r"(\d{2,4})\s*[*xX×]\s*(\d{2,4})", value or "")
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def parse_thickness(value: str) -> int | None:
    match = re.search(r"(\d+(?:\.\d+)?)\s*T", value or "", re.I)
    if not match:
        return None
    return int(float(match.group(1)))


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").upper())


def format_decimal(value: float) -> str:
    if math.isclose(value, round(value)):
        return str(int(round(value)))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def is_snt_product(product: dict) -> bool:
    values = [
        product.get("majorCategory"),
        product.get("catalogSource"),
        product.get("maker"),
        product.get("kind"),
        product.get("managementCode"),
        product.get("id"),
    ]
    return any(str(value or "").upper().startswith("SNT") or str(value or "").lower().startswith("snt-") for value in values)


def read_json_array(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8") as file:
        value = json.load(file)
    if not isinstance(value, list):
        raise ValueError(f"JSON 배열이 아닙니다: {path}")
    return value


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f"{path.name}.{datetime.now().strftime('%Y%m%d%H%M%S%f')}.tmp")
    with temp_path.open("w", encoding="utf-8") as file:
        json.dump(value, file, ensure_ascii=False, indent=2)
        file.write("\n")
    temp_path.replace(path)


if __name__ == "__main__":
    main()
